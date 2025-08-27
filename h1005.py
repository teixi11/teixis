from __future__ import annotations

from flask import Flask, request, render_template_string, redirect, url_for
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
import os, re, datetime
from unicodedata import normalize as ucnorm
from difflib import SequenceMatcher  # fuzzy match

# File-lock (opcional)
try:
    import portalocker
    HAVE_PORTALOCKER = True
except Exception:
    HAVE_PORTALOCKER = False

app = Flask(__name__)
app.secret_key = "dev-key"
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

# ========= Config =========
BASE_TEMPLATE_DIR = os.environ.get(
    "QS_BASE_DIR",
    r"C:\Hola"
)
ALLOWED_EXT = {".xlsx"}

# --- Centro de cost: Excel maestro (ruta fija C:\Hola) ---
DEFAULT_CC_PATH = r"C:\Hola\CC.xlsx"
CC_XLSX_PATH = os.environ.get("QS_CC_PATH", DEFAULT_CC_PATH)

# ========= Utilidades =========
def listar_excels_recursivo(base_dir: str) -> list[str]:
    res = []
    if not os.path.isdir(base_dir):
        return res
    for root, _, files in os.walk(base_dir):
        rlow = root.lower()
        if "qs" not in rlow:
            continue
        if "fitxes" in rlow:
            continue
        for f in files:
            if os.path.splitext(f)[1].lower() in ALLOWED_EXT:
                rel = os.path.relpath(os.path.join(root, f), base_dir).replace("\\", "/")
                res.append(rel)
    res.sort()
    return res

def norm_txt(s: str) -> str:
    """Minúsculas, sin acentos, limpia signos, tolera ':' y '/' y colapsa espacios."""
    if not isinstance(s, str):
        return ""
    s = ucnorm("NFKD", s)
    s = "".join(ch for ch in s if ord(ch) < 128)
    s = s.lower()
    s = s.replace(":", " ").replace("%", " % ").replace("(", " ").replace(")", " ").replace("/", " / ")
    s = re.sub(r"[^a-z0-9%/.\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def piece_name_from_relpath(relpath: str) -> str:
    base = os.path.basename(relpath)
    name, _ = os.path.splitext(base)
    return name

# --- merged helpers (compat) ---
def merged_range(ws, row: int, col: int):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None

def merged_top_left(ws, row: int, col: int):
    rng = merged_range(ws, row, col)
    if rng:
        return rng.min_row, rng.min_col
    return row, col

def read_cell_value(ws, row: int, col: int):
    r0, c0 = merged_top_left(ws, row, col)
    return ws.cell(r0, c0).value

def a1_ref(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"

def cell_ref_for_write(ws, row: int, col: int) -> str:
    r0, c0 = merged_top_left(ws, row, col)
    return a1_ref(r0, c0)

def read_value_a1(ws, a1: str):
    m = re.match(r"([A-Za-z]+)(\d+)$", a1)
    if not m:
        return None
    col_str, row_str = m.group(1), m.group(2)
    col = column_index_from_string(col_str.upper())
    row = int(row_str)
    r0, c0 = merged_top_left(ws, row, col)
    return ws.cell(r0, c0).value

# ========= Utilidades merges (con cache) =========
def build_merge_map(ws):
    m = {}
    for rng in ws.merged_cells.ranges:
        tl = (rng.min_row, rng.min_col)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                m[(r, c)] = tl
    return m

def merged_top_left_cached(merge_map, row: int, col: int):
    return merge_map.get((row, col), (row, col))

def read_cell_value_cached(ws, merge_map, row: int, col: int):
    r0, c0 = merged_top_left_cached(merge_map, row, col)
    return ws.cell(r0, c0).value

def cell_ref_for_write_cached(merge_map, row: int, col: int) -> str:
    r0, c0 = merged_top_left_cached(merge_map, row, col)
    return a1_ref(r0, c0)

def read_value_a1_cached(ws, merge_map, a1: str):
    m = re.match(r"([A-Za-z]+)(\d+)$", a1)
    if not m:
        return None
    col_str, row_str = m.group(1), m.group(2)
    col = column_index_from_string(col_str.upper())
    row = int(row_str)
    r0, c0 = merged_top_left_cached(merge_map, row, col)
    return ws.cell(r0, c0).value

def label_matches(text: str, variants: list[str]) -> bool:
    """Coincidencia flexible para rótulos."""
    t = norm_txt(text).strip(" /")
    t_short = t[:64]
    for v in variants:
        nv = norm_txt(v).strip(" /")
        if t.startswith(nv) or f" {nv} " in f" {t} ":
            return True
        if SequenceMatcher(None, t_short, nv[:64]).ratio() >= 0.80:
            return True
    return False

# ========= Reglas por fase =========
PHASE_RULES = {
    "Torn": {
        "Segons/peça": {
            "variants": ["segons/peça", "segons / peça", "s/peça", "segons peça", "segons-pesa", "segundos por pieza", "s por pieza"],
            "target": lambda r, c: (r + 1, c),
        }
    },
    "Neteja": {
        "Temps cicle (min)": {"variants": ["temps cicle(min)", "temps cicle (min)", "tiempo ciclo", "tiempo de ciclo", "temps min", "temps minut"], "target": lambda r, c: (r, c + 2)},
        "Quantitat": {"variants": ["quantitat", "cantidad", "qt"], "target": lambda r, c: (r, c + 2)},
    },
    "Vibrar": {
        "Temps cicle (min)": {"variants": ["temps cicle(min)", "temps cicle (min)", "tiempo ciclo", "temps min", "temps minut"], "target": lambda r, c: (r, c + 2)},
        "Quantitat": {"variants": ["quantitat", "cantidad", "qt"], "target": lambda r, c: (r, c + 2)},
    },
    "SubZero": {"kg/1000": {"variants": ["kg/1000","kg / 1000","kg 1000","kg por 1000","kg/1000 pcs","kg / 1000 pcs"], "target": lambda r, c: (r, c + 2)}},
    "Trempat": {"kg/1000": {"variants": ["kg/1000","kg / 1000","kg 1000","kg por 1000","kg/1000 pcs","kg / 1000 pcs"], "target": lambda r, c: (r, c + 2)}},
    "Revindre": {"kg/1000": {"variants": ["kg/1000","kg / 1000","kg 1000","kg por 1000","kg/1000 pcs","kg / 1000 pcs"], "target": lambda r, c: (r, c + 2)}},
    "Decapar": {"kg/1000": {"variants": ["kg/1000","kg / 1000","kg 1000","kg por 1000","kg/1000 pcs","kg / 1000 pcs"], "target": lambda r, c: (r, c + 2)}},
    "Centres exteriors": {"€/1000": {"variants": ["€/1000","eur/1000","eur 1000","euros/1000","€ / 1000"], "target": lambda r, c: (r, c + 2)}},
    "Desbarbat químic": {
        "Gruix de capa en radi que es menja (centèssimes)": {
            "variants": ["gruix de capa en radi que es menja (centèssimes)","gruix de capa en radi que es menja (centessimes)","espesor capa centesimas","espesor capa"],
            "target": lambda r, c: (r, c + 4),
        }
    },
    "Rectificar barres": {
        "Segons / barra": {"variants": ["segons / barra","segons/barra","s/barra","segundos por barra"], "target": lambda r, c: (r, c + 2)},
        "Peces / barra": {"variants": ["peces / barra","piezas / barra"], "target": lambda r, c: (r, c + 2)},
    },
    "Màq visió 100%": {"Segons/peça": {"variants": ["segons/peça","segons / peça","s/peça","segons peça","segundos por pieza"], "target": lambda r, c: (r, c + 2)}},
    "Rectificar peça": {"Segons/peça": {"variants": ["segons/peça","segons / peça","s/peça","segons peça","segundos por pieza"], "target": lambda r, c: (r, c + 4)}},
    "Mostreig Visual": {
        "Temps (min)": {"variants": ["temps (min)","tiempo (min)","temps min","temps minut","temps"], "target": lambda r, c: (r, c + 2)},
        "Quantitat":   {"variants": ["quantitat","cantidad","qt"], "target": lambda r, c: (r, c + 2)},
    },
    "Embalar manualment": {
        "Temps (min)": {"variants": ["temps (min)","tiempo (min)","temps minut","temps"], "target": lambda r, c: (r, c + 2)},
        "Quantitat":   {"variants": ["quantitat","cantidad","qt"], "target": lambda r, c: (r, c + 2)},
    },
    "Embalar màq. Visió": {
        "Temps (min)": {"variants": ["temps (min)","tiempo (min)","temps min","temps minut","temps"], "target": lambda r, c: (r, c + 2)},
        "Quantitat":   {"variants": ["quantitat","cantidad","qt"], "target": lambda r, c: (r, c + 2)},
    },
}

# ========= Carga CC.xlsx (A,B,C + D,E) con caché =========
_CC_CACHE = {"mtime": None, "data": None}  # data: types, items_by_type, by_code, de_pairs, neteja_matrix

def load_cc_table():
    path = CC_XLSX_PATH
    if not os.path.isfile(path):
        return {"types": [], "items_by_type": {}, "by_code": {}, "de_pairs": [], "neteja_matrix": {"machines": [], "programs": [], "values": {}}}

    mtime = os.path.getmtime(path)
    if _CC_CACHE["mtime"] == mtime and _CC_CACHE["data"] is not None:
        return _CC_CACHE["data"]

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    types = set()
    items_by_type = {}
    by_code = {}
    de_pairs = []  # [{num:D, text:E}]

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        codi = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        desc = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        tipus = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        numd = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        etxt = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""

        if codi and desc:
            types.add(tipus)
            items_by_type.setdefault(tipus, []).append({"code": codi, "desc": desc})
            by_code[codi] = {"desc": desc, "tipus": tipus}

        if numd or etxt:
            de_pairs.append({"num": numd, "text": etxt})

    # dedupe por descripción dentro de cada tipus
    for t, lst in list(items_by_type.items()):
        seen = set()
        dedup = []
        for it in lst:
            key = norm_txt(it["desc"])
            if key in seen:
                continue
            seen.add(key)
            dedup.append(it)
        dedup.sort(key=lambda x: x["desc"].lower())
        items_by_type[t] = dedup

    types_sorted = sorted(types, key=lambda s: (s == "", s.lower() if isinstance(s, str) else str(s)))

    # ========= Neteja MATRIX: L2:L10 máquinas, M1:R1 programas, intersección L1:R10 =========
    try:
        col_L = column_index_from_string('L')
        col_M = column_index_from_string('M')
        col_R = column_index_from_string('R')

        machines = []
        for r in range(2, 11):  # L2..L10
            val = ws.cell(row=r, column=col_L).value
            if val is None:
                continue
            txt = str(val).strip()
            if txt:
                machines.append(txt)

        programs = []
        for c in range(col_M, col_R + 1):  # M1..R1
            val = ws.cell(row=1, column=c).value
            if val is None:
                continue
            txt = str(val).strip()
            if txt:
                programs.append(txt)

        values = {}
        for idx_r, r in enumerate(range(2, 2 + len(machines))):
            mname = machines[idx_r]
            rowvals = {}
            for idx_c, c in enumerate(range(col_M, col_M + len(programs))):
                pname = programs[idx_c]
                v = ws.cell(row=r, column=c).value
                t = None
                if isinstance(v, (int, float)):
                    t = float(v)
                else:
                    try:
                        t = float(str(v).replace(",", "."))
                    except Exception:
                        t = None
                rowvals[pname] = t
            values[mname] = rowvals

        neteja_matrix = {"machines": machines, "programs": programs, "values": values}
    except Exception:
        neteja_matrix = {"machines": [], "programs": [], "values": {}}

    data = {
        "types": types_sorted,
        "items_by_type": items_by_type,
        "by_code": by_code,
        "de_pairs": de_pairs,
        "neteja_matrix": neteja_matrix,
    }
    _CC_CACHE["mtime"] = mtime
    _CC_CACHE["data"] = data
    wb.close()
    return data

# ========= Detección de fases + campos =========
def detect_schema(abs_path: str, debug: bool = False):
    wb = load_workbook(abs_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    merge_map = build_merge_map(ws)
    min_c, min_r, max_c, max_r = range_boundaries(ws.calculate_dimension())

    # detectar filas con "Fase"
    phase_rows = []
    for r in range(min_r, max_r + 1):
        for c in range(min_c, min(min_c + 3, max_c) + 1):
            val = read_cell_value_cached(ws, merge_map, r, c)
            if isinstance(val, str):
                txt = norm_txt(val)
                if " fase " in f" {txt} " or txt.startswith("fase"):
                    phase_rows.append((r, val.strip()))
                    break

    phases_out = []
    for i, (row, title) in enumerate(phase_rows):
        next_row = phase_rows[i + 1][0] if i + 1 < len(phase_rows) else (max_r + 1)
        block_start = row + 1
        block_end = max(block_start, next_row - 1)

        title_norm = norm_txt(title)
        if   "torn" in title_norm:                                   phase_key = "Torn"
        elif "neteja" in title_norm:                                  phase_key = "Neteja"
        elif "vibr" in title_norm:                                    phase_key = "Vibrar"
        elif "subzero" in title_norm:                                 phase_key = "SubZero"
        elif "trempat" in title_norm:                                 phase_key = "Trempat"
        elif "revindre" in title_norm:                                phase_key = "Revindre"
        elif "decapar" in title_norm:                                 phase_key = "Decapar"
        elif "centres exteriors" in title_norm or "centros exteriores" in title_norm:
                                                                       phase_key = "Centres exteriors"
        elif "desbarbat" in title_norm:                               phase_key = "Desbarbat químic"
        elif "rectificar barres" in title_norm:                       phase_key = "Rectificar barres"
        elif "maq visio" in title_norm or "màq visió" in title_norm or "maq vision" in title_norm or "maq visión" in title_norm:
                                                                       phase_key = "Màq visió 100%"
        elif "rectificar peça" in title_norm or "rectificar pieza" in title_norm:
                                                                       phase_key = "Rectificar peça"
        elif "mostreig visual" in title_norm:                         phase_key = "Mostreig Visual"
        elif "embalar" in title_norm and ("maq" in title_norm or "visio" in title_norm or "visión" in title_norm):
                                                                       phase_key = "Embalar màq. Visió"
        elif "embalar" in title_norm:                                 phase_key = "Embalar manualment"
        else:
            if debug:
                print(f"[detect_schema] Fase desconocida: {title!r}")
            continue

        rules = PHASE_RULES.get(phase_key, {})
        if not rules:
            continue

        detected_fields = []
        obs_field = None
        cc_field = None
        quant_cell = None  # para Neteja

        # buscar campos por bloque
        for rr in range(block_start, block_end + 1):
            for cc in range(min_c, max_c + 1):
                raw = read_cell_value_cached(ws, merge_map, rr, cc)

                # Observacions
                if isinstance(raw, str) and (obs_field is None):
                    t = norm_txt(raw)
                    if re.search(r'\bobservaci(ó|o)nes?\b|\bobservacions?\b|\bobs\.?\b', t):
                        r0, c0 = merged_top_left_cached(merge_map, rr, cc)
                        tr, tc = r0, c0 + 1
                        tr = min(max(tr, min_r), max_r)
                        tc = min(max(tc, min_c), max_c)
                        a1 = cell_ref_for_write_cached(merge_map, tr, tc)
                        cur = read_value_a1_cached(ws, merge_map, a1)
                        obs_field = {"label": "Observacions", "cell": a1, "value": "" if cur is None else str(cur)}
                        continue

                # Centre de cost
                if isinstance(raw, str) and (cc_field is None):
                    t = norm_txt(raw)
                    if re.search(r'\bcentre de cost\b|\bcentro de coste\b|\bcentro de costo\b|\bcc\b', t):
                        r0, c0 = merged_top_left_cached(merge_map, rr, cc)
                        tr, tc = r0, c0 + 1
                        tr = min(max(tr, min_r), max_r)
                        tc = min(max(tc, min_c), max_c)
                        a1 = cell_ref_for_write_cached(merge_map, tr, tc)
                        cur = read_value_a1_cached(ws, merge_map, a1)
                        cc_field = {"label": "Centre de cost", "cell": a1, "value": "" if cur is None else str(cur)}
                        continue

                if not isinstance(raw, str):
                    continue

                for nice_label, rule in rules.items():
                    if label_matches(raw, rule["variants"]):
                        r0, c0 = merged_top_left_cached(merge_map, rr, cc)
                        tr, tc = rule["target"](r0, c0)
                        tr = min(max(tr, min_r), max_r)
                        tc = min(max(tc, min_c), max_c)
                        a1 = cell_ref_for_write_cached(merge_map, tr, tc)
                        cur = read_value_a1_cached(ws, merge_map, a1)

                        if (phase_key == "Neteja") and (nice_label == "Quantitat"):
                            quant_cell = a1
                            continue

                        if not any(f["label"] == nice_label for f in detected_fields):
                            detected_fields.append({"label": nice_label, "cell": a1, "value": "" if cur is None else str(cur)})
                        break

        # extras Torn
        torn_extra_times = []
        torn_obs_nums = []
        torn_base_cell = ""

        if phase_key == "Torn":
            base_cell = next((f["cell"] for f in detected_fields if f["label"]=="Segons/peça"), "")
            torn_base_cell = base_cell
            m = re.match(r"([A-Za-z]+)(\d+)$", base_cell or "")
            if m:
                base_col = column_index_from_string(m.group(1))
                base_row = int(m.group(2))
                for off in range(1, 21):
                    v = ws.cell(base_row + off, base_col).value
                    if v is None or (isinstance(v, str) and str(v).strip()==""):
                        break
                    torn_extra_times.append(str(v))

            if obs_field and obs_field.get("value"):
                obs = obs_field["value"]
                mline = re.search(r"(?mi)^\s*Torns\s*:\s*(.+)$", obs or "")
                if mline:
                    payload = mline.group(1)
                    nums = re.findall(r"torn\s*([0-9]+)", payload, flags=re.I)
                    torn_obs_nums = [n.strip() for n in nums if n.strip()]

        if cc_field is None:
            cc_field = {"label": "Centre de cost", "cell": "", "value": ""}

        ordered_fields = []
        for fld in ([cc_field] + ([obs_field] if obs_field else []) + detected_fields):
            if fld and not any(x["label"] == fld["label"] for x in ordered_fields):
                ordered_fields.append(fld)

        phase_obj = {"title": title, "phase_key": phase_key, "fields": ordered_fields}
        if phase_key == "Neteja":
            phase_obj["quant_cell"] = quant_cell or ""
            phase_obj["obs_cell"] = (obs_field["cell"] if obs_field else "")
        if phase_key == "Torn":
            if obs_field: phase_obj["obs_cell"] = obs_field["cell"]
            phase_obj["torn_base_cell"] = torn_base_cell
            phase_obj["torn_extra_times"] = torn_extra_times
            phase_obj["torn_obs_nums"] = torn_obs_nums

        phases_out.append(phase_obj)

    sheet_name = ws.title
    wb.close()

    # fusionar fases con mismo título
    groups = {}
    for ph in phases_out:
        key = norm_txt(ph["title"])
        groups.setdefault(key, []).append(ph)

    coalesced = []
    for _, items in groups.items():
        if len(items) == 1:
            coalesced.append(items[0])
            continue

        rep = max(items, key=lambda p: len(p.get("fields", [])))
        seen = set()
        merged_fields = []
        for f in rep.get("fields", []):
            if f["label"] not in seen:
                merged_fields.append(f); seen.add(f["label"])
        for it in items:
            for f in it.get("fields", []):
                if f["label"] not in seen:
                    merged_fields.append(f); seen.add(f["label"])

        out = {"title": rep["title"], "phase_key": rep["phase_key"], "fields": merged_fields}
        if rep["phase_key"] == "Neteja":
            qc = next((ph.get("quant_cell") for ph in items if ph.get("quant_cell")), "")
            oc = next((ph.get("obs_cell") for ph in items if ph.get("obs_cell")), "")
            out["quant_cell"] = qc
            out["obs_cell"] = oc
        if rep["phase_key"] == "Torn":
            oc = next((ph.get("obs_cell") for ph in items if ph.get("obs_cell")), "")
            if oc: out["obs_cell"] = oc
            t_base = next((ph.get("torn_base_cell") for ph in items if ph.get("torn_base_cell")), "")
            t_times = next((ph.get("torn_extra_times") for ph in items if ph.get("torn_extra_times")), [])
            t_nums = next((ph.get("torn_obs_nums") for ph in items if ph.get("torn_obs_nums")), [])
            out["torn_base_cell"] = t_base
            out["torn_extra_times"] = t_times
            out["torn_obs_nums"] = t_nums

        coalesced.append(out)

    return {"sheet": sheet_name, "phases": coalesced}

# ========= HTML / CSS =========
BASE_CSS = """
:root{
  --bg1:#0f1020; --bg2:#0f1a3a; --card:#121528cc; --ink:#E8EAED;
  --accent1:#7c3aed; --accent2:#00e5ff; --ok:#16a34a; --muted:#9aa0a6;
  --radius: 18px; --phase-gap: clamp(16px, 2.4vw, 32px);
}
*{box-sizing:border-box}
html,body{height:100%}
body{
  margin:0; color:var(--ink); background: linear-gradient(135deg,var(--bg1),var(--bg2));
  font-family: Arial; background-attachment: fixed;
}
/* Hero */
.hero{ position:relative; padding:clamp(24px,3vw,36px); }
.brand{ display:flex; align-items:center; gap:12px; }
.brand .logo-dot{
  width:14px; height:14px; border-radius:50%;
  background: radial-gradient(circle at 30% 30%, var(--accent2), var(--accent1));
  box-shadow: 0 0 18px rgba(0,229,255,.6), 0 0 28px rgba(124,58,237,.35);
  animation: pulse 2.4s ease-in-out infinite;
}
@keyframes pulse{ 0%,100%{ transform: scale(1);} 50%{ transform: scale(1.2);} }
.h1{ font-weight:800; letter-spacing:-0.02em; font-size:clamp(20px, 2.6vw, 28px); }
/* Gyro */
.gyro-wrap{ position:absolute; right: clamp(10px,3vw,40px); top: clamp(6px,1vw,10px); width: 96px; height: 96px; perspective: 1000px; pointer-events:none;}
.gyro{ position:relative; width:100%; height:100%; transform-style:preserve-3d; animation: gyroSpin 22s linear infinite; }
@keyframes gyroSpin{ from{ transform: rotateY(0deg);} to{ transform: rotateY(360deg);} }
.ring{ position:absolute; inset:0; border-radius:50%; border: 2px solid rgba(255,255,255,.35);
  box-shadow: 0 0 18px rgba(124,58,237,.25), inset 0 0 18px rgba(0,229,255,.18); }
.ring.x{ transform: rotateX(72deg); animation: spinX 10s linear infinite; }
.ring.y{ transform: rotateY(72deg); animation: spinY 14s linear infinite reverse; }
.ring.z{ transform: rotateZ(0);  animation: spinZ 18s linear infinite; }
@keyframes spinX{ from{transform: rotateX(72deg) rotateZ(0);} to { transform: rotateX(72deg) rotateZ(360deg);} }
@keyframes spinY{ from{transform: rotateY(72deg) rotateZ(0);} to { transform: rotateY(72deg) rotateZ(360deg);} }
@keyframes spinZ{ from{transform: rotateZ(0);} to { transform: rotateZ(360deg);} }
/* Container */
.container-slim{ max-width: 980px; margin: 0 auto; padding: 0 clamp(16px,2vw,24px) 56px; }
/* Card */
.card{
  background: linear-gradient(180deg, rgba(255,255,255,.06), rgba(255,255,255,.03));
  border: 1px solid rgba(255,255,255,.10);
  backdrop-filter: blur(10px);
  border-radius: var(--radius);
  box-shadow: 0 12px 50px rgba(0,0,0,.35);
}
/* Form */
.label{ font-size: 13px; color:#ffffff; font-weight: 450; margin-bottom: 8px; display:flex; align-items:center; gap:4px;}
.label svg{ width:14px; height:14px; vertical-align:text-bottom; opacity:.9; }
.form-select, .form-input, textarea{
  width: 100%; background:#0e1224; color:var(--ink); border: 1px solid #2a2f4a; border-radius: 12px;
  padding: 12px 14px; outline: none; transition: border-color .2s, box-shadow .2s, transform .06s;
}
.form-select:focus, .form-input:focus, textarea:focus{ border-color: var(--accent2); box-shadow: 0 0 0 3px rgba(0,229,255,.18); }
.form-input:active{ transform: scale(.998); }
/* Buttons */
.btn, .btn-link{ display:inline-flex; align-items:center; }
.btn{
  cursor:pointer; user-select:none; border:0; border-radius:14px; padding:12px 16px; font-weight:800; color:#0b0f1e;
  background: linear-gradient(90deg, var(--accent1), var(--accent2));
  box-shadow: 0 6px 24px rgba(0,229,255,.28), 0 6px 20px rgba(124,58,237,.20);
  transition: transform .08s ease, filter .2s ease;
}
.btn:hover{ filter: brightness(1.08) saturate(1.05);}
.btn:active{ transform: translateY(1px); }
.btn-link{ background: transparent; border: none; color: #a6ecff; padding: 0; font-weight: 300; }
.btn-link:hover{ text-decoration: underline; }
/* Sticky footer */
.sticky-footer{ position: sticky; bottom: 0; background: #0b0f1e; padding: 12px;
  border: 1px solid rgba(255,255,255,.08); border-radius: 14px;}
/* Phase chip (gradiente azul como botón) */
.phase-chip{
  display:inline-flex; align-items:center; gap:10px; padding:8px 12px; border-radius:12px;
  background: linear-gradient(90deg, var(--accent1), var(--accent2));
  color:#0b0f1e; font-weight:800;
  box-shadow: 0 6px 24px rgba(0,229,255,.28), 0 6px 20px rgba(124,58,237,.20);
  position:relative; overflow:hidden;
}
.phase-chip:before{
  content:""; position:absolute; inset:-2px; border-radius:12px;
  background: linear-gradient(90deg, rgba(124,58,237,.0), rgba(0,229,255,.35), rgba(124,58,237,.0));
  filter: blur(6px); animation: chipGlow 2.2s linear infinite;
}
@keyframes chipGlow{
  0%{ transform: translateX(-120%); opacity:.6;}
  50%{ opacity:1;}
  100%{ transform: translateX(120%); opacity:.6;}
}
.phase-head{ display:flex; align-items:center; gap:10px; margin-bottom:8px; }
/* Grid & helpers */
.actions{ display:flex; gap:10px; flex-wrap:nowrap; align-items:center; }
@media (max-width:560px){ .actions{ flex-wrap:wrap; } }
.muted{ color: var(--muted); }
.grid{ display:grid; gap:16px; }
.grid-2{ grid-template-columns: 1fr; }
@media (min-width: 840px){ .grid-2{ grid-template-columns: 1fr 1fr; }}
/* Alert */
.alert{ padding: 10px 12px; border:1px solid #8b1d1d; background:#2a0b0b; color:#ffd3d3; border-radius: 12px; }
/* Loader overlay */
#loading-overlay{ position:fixed; inset:0; display:none; align-items:center; justify-content:center; z-index:9999;
  background: radial-gradient( circle at 30% 20%, rgba(124,58,237,.18), transparent 60% ),
              radial-gradient( circle at 70% 70%, rgba(0,229,255,.15), transparent 55% ),
              rgba(10,12,24,.82); }
.loader-box{ min-width:280px; padding:18px; border-radius:16px; text-align:center;
  background: linear-gradient(180deg, rgba(255,255,255,.10), rgba(255,255,255,.06));
  border: 1px solid rgba(255,255,255,.18); box-shadow: 0 20px 60px rgba(0,0,0,.45); }
.spinner{ width:40px; height:40px; border-radius:50%;
  border: 4px solid rgba(255,255,255,.18); border-top-color: var(--accent2);
  animation: spin 1.1s linear infinite; margin: 0 auto; }
@keyframes spin{ to{ transform: rotate(360deg);} }
/* Phase spacing */
.phase-card{ border-radius: 16px; border:1px solid rgba(255,255,255,.10); }
.phase-card + .phase-card{ margin-top: var(--phase-gap); }
.phase-card:last-of-type{ margin-bottom: var(--phase-gap); }
/* Reduce motion */
@media (prefers-reduced-motion: reduce) {
  .gyro{ animation-duration: 28s !important; }
  .ring.x, .ring.y, .ring.z { animation-duration: 16s !important; }
  .phase-chip:before{ animation-duration: 4s !important; }
}
/* Torn minis */
.inline-compact{ display:flex; gap:8px; align-items:center; }
.mini-select{ min-width:84px; width:auto; padding:6px 8px; border-radius:8px; }
.mini-input{ max-width:160px; }
"""

INDEX_HTML = """
<!doctype html>
<html lang="es">
  <head>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>QS</title>
    <link href="https://fonts.googleapis.com/css2?family=Carlito:wght@300&display=swap" rel="stylesheet">
    <style>{{ BASE_CSS }}</style>
    <script src="https://unpkg.com/lucide@latest"></script>
  </head>
  <body>
    <div id="loading-overlay">
      <div class="loader-box">
        <div class="spinner" aria-label="Cargando"></div>
        <div class="mt-2" style="margin-top:10px; font-weight:300;">Cargando…</div>
      </div>
    </div>

    <header class="hero">
      <div class="brand">
        <div class="logo-dot"></div>
        <div class="h1">QS</div>
      </div>
      <div class="gyro-wrap" aria-hidden="true">
        <div class="gyro">
          <div class="ring x"></div>
          <div class="ring y"></div>
          <div class="ring z"></div>
        </div>
      </div>
    </header>

    <main class="container-slim grid" style="gap:20px;">
      {% if msg %}
        <div class="alert" role="alert">{{ msg }}</div>
      {% endif %}
      <div class="card" style="padding:18px;">
        <form id="index-form" method="get" action="{{ url_for('edit') }}">
          <div class="grid grid-2">
            <div>
              <div class="label">Selecciona Excel del PC (busca en subcarpetas)</div>
              <select class="form-select" name="file" required>
                <option value="" disabled selected>-- Elige un archivo --</option>
                {% for f in files %}<option value="{{ f }}">{{ f }}</option>{% endfor %}
              </select>
              <div class="muted" style="margin-top:6px; font-size:12px;">Carpeta raíz: <code>{{ base_dir }}</code></div>
            </div>
          </div>
          <div class="actions" style="margin-top:16px;">
            <button class="btn" type="submit">Detectar y continuar</button>
            <button class="btn-link" type="button" onclick="location.reload()">Recargar</button>
          </div>
        </form>
      </div>
    </main>

    <script>
      (function () {
        const overlay = document.getElementById('loading-overlay');
        const form = document.getElementById('index-form');
        if (!overlay || !form) return;
        form.addEventListener('submit', function () {
          overlay.style.display = 'flex';
        });
        window.addEventListener('pagehide', () => {
          overlay.style.display = 'none';
        }, { once: true });
      })();
      lucide.createIcons();
    </script>
  </body>
</html>
"""

EDIT_HTML = """
<!doctype html>
<html lang="es">
  <head>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>QS – {{ piece_name }}</title>
    <link href="https://fonts.googleapis.com/css2?family=Carlito:wght@300&display=swap" rel="stylesheet">
    <style>{{ BASE_CSS }}</style>
    <script src="https://unpkg.com/lucide@latest"></script>
  </head>
  <body>
    <div id="loading-overlay">
      <div class="loader-box">
        <div class="spinner" aria-label="Guardando"></div>
        <div class="mt-2" style="margin-top:10px; font-weight:300;">Guardando…</div>
      </div>
    </div>

    <header class="hero">
      <div class="brand">
        <div class="logo-dot"></div>
        <div class="h1">QS – {{ piece_name }}</div>
      </div>
      <div class="gyro-wrap" aria-hidden="true">
        <div class="gyro">
          <div class="ring x"></div>
          <div class="ring y"></div>
          <div class="ring z"></div>
        </div>
      </div>
    </header>

    <main class="container-slim grid" style="gap:18px;">
      {% if not schema.phases %}
        <div class="card" style="padding:18px;">
          <div class="muted">No se detectaron campos en las fases configuradas.</div>
          <div class="actions" style="margin-top:10px;">
            <a class="btn-link" href="{{ url_for('index') }}">Volver</a>
          </div>
        </div>
      {% else %}
        <form id="edit-form" method="post" action="{{ url_for('edit') }}">
          <input type="hidden" name="file" value="{{ relpath }}">

          {% for p in schema.phases %}
            {% set pidx = loop.index0 %}
            <div class="card phase-card" style="padding:16px;">
              <div class="phase-head">
                <div class="phase-chip">
                  <span>{{ p.title }}</span>
                </div>
                <div class="badge">{{ p.phase_key }}</div>
              </div>

              {# Campos especiales Neteja #}
              {% if p.phase_key == "Neteja" %}
                {% if p.quant_cell %}<input type="hidden" name="meta_quant_cell__{{ pidx }}" value="{{ p.quant_cell }}">{% endif %}
                {% if p.obs_cell %}<input type="hidden" name="meta_obs_cell__{{ pidx }}" value="{{ p.obs_cell }}">{% endif %}

                <div class="grid grid-2" style="margin-top:6px;">
                  <div data-field-group>
                    <div class="label">Tipus bastidor <i data-lucide="clipboard"></i></div>
                    <input class="form-input" type="text" name="bast_tipus__{{ pidx }}" value="{{ sticky.get('bast_tipus__' ~ pidx, '') }}">
                  </div>
                  <div data-field-group>
                    <div class="label">Nº bastidors <i data-lucide="package"></i></div>
                    <input class="form-input" type="number" inputmode="numeric" name="bast_num__{{ pidx }}" value="{{ sticky.get('bast_num__' ~ pidx, '') }}">
                  </div>
                  <div data-field-group>
                    <div class="label">Peces per bastidor <i data-lucide="package"></i></div>
                    <input class="form-input" type="number" inputmode="numeric" name="bast_ppb__{{ pidx }}" value="{{ sticky.get('bast_ppb__' ~ pidx, '') }}">
                  </div>
                  <div data-field-group>
                    <div class="label">Nº màquina (neteja) <i data-lucide="settings"></i></div>
                    <select class="form-select neteja-maq" name="neteja_maq__{{ pidx }}" data-selected="{{ sticky.get('neteja_maq__' ~ pidx, '') }}">
                      <option value="">-- Selecciona --</option>
                    </select>
                  </div>
                  <div data-field-group>
                    <div class="label">Programa <i data-lucide="timer"></i></div>
                    <select class="form-select neteja-prog" name="neteja_prog__{{ pidx }}" data-selected="{{ sticky.get('neteja_prog__' ~ pidx, '') }}">
                      <option value="">-- Selecciona --</option>
                    </select>
                  </div>
                  <div data-field-group>
                    <div class="label">Nau <i data-lucide="building"></i></div>
                    <select class="form-select" name="neteja_nau__{{ pidx }}">
                      {% set want = sticky.get('neteja_nau__' ~ pidx, '') %}
                      <option value="">-- Selecciona --</option>
                      {% for n in ['N0','N1','N2','N3'] %}
                        <option value="{{ n }}" {% if want==n %}selected{% endif %}>{{ n }}</option>
                      {% endfor %}
                    </select>
                  </div>
                </div>
              {% endif %}

              {% if p.phase_key == "Torn" and p.obs_cell %}
                <input type="hidden" name="meta_obs_cell__{{ pidx }}" value="{{ p.obs_cell }}">
              {% endif %}

              <div class="grid grid-2" style="margin-top:6px;">
                {% for f in p.fields %}
                  {% set fidx = loop.index0 %}
                  {% set label_l = f.label|lower %}
                  {% set is_obs = ('observ' in label_l) %}
                  {% set is_cc = (f.label == "Centre de cost") %}
                  {% set is_torn_seg = (p.phase_key == "Torn" and f.label == "Segons/peça") %}
                  {% set numeric_labels = ["Segons/peça","Temps cicle (min)","Quantitat","kg/1000","€/1000","Segons / barra","Peces / barra","Temps (min)"] %}
                  {% set is_numeric = (numeric_labels|join('||')) and (not is_obs) and (f.label in numeric_labels) %}

                  <div data-field-group style="{% if is_obs %}grid-column: 1 / -1;{% endif %}">
                    <div class="label">
                      {{ f.label }}
                      {% if 'temps' in label_l or 'segons' in label_l %}<i data-lucide="clock"></i>{% endif %}
                      {% if 'quantitat' in label_l or 'peces / barra' in label_l or 'peces/' in label_l %}<i data-lucide="package"></i>{% endif %}
                      {% if 'kg/1000' in label_l %}<i data-lucide="scale-3d"></i>{% endif %}
                      {% if '€/1000' in label_l %}<i data-lucide="euro"></i>{% endif %}
                      {% if 'longitud' in label_l %}<i data-lucide="ruler"></i>{% endif %}
                      {% if is_obs %}<i data-lucide="clipboard-list"></i>{% endif %}
                    </div>

                    {% if is_cc %}
                      {# CC fijo / oculto en Neteja y Centres exteriors; normal en resto #}
                      {% if p.phase_key == "Neteja" or p.phase_key == "Centres exteriors" %}
                        <input type="hidden" name="field__{{ fidx }}__{{ pidx }}" value="{% if p.phase_key == 'Neteja' %}31050{% else %}39000{% endif %}">
                        <input type="hidden" name="meta_label__{{ fidx }}__{{ pidx }}" value="{{ f.label }}">
                        <input type="hidden" name="meta_cell__{{ fidx }}__{{ pidx }}" value="{{ f.cell }}">
                      {% else %}
                        <div style="display:flex; gap:8px; flex-wrap:wrap;">
                          <select class="form-select cc-tipus" data-target="tipus"></select>
                          <select class="form-select cc-maquina" data-target="maquina"></select>
                        </div>
                        <div style="display:flex; gap:8px; flex-wrap:wrap; margin-top:8px;">
                          <select class="form-select cc-num" data-target="num">
                            <option value="">-- Nº Màquina (col D) — por E --</option>
                          </select>
                        </div>
                        <input type="hidden" name="field__{{ fidx }}__{{ pidx }}" class="cc-code" value="{{ f.value }}">
                        <input type="hidden" name="meta_label__{{ fidx }}__{{ pidx }}" value="{{ f.label }}">
                        <input type="hidden" name="meta_cell__{{ fidx }}__{{ pidx }}" value="{{ f.cell }}">
                        <div class="actions" style="margin-top:6px;">
                          <button class="btn-link" type="button" onclick="ccVaciar(this)">Vaciar</button>
                        </div>
                      {% endif %}

                    {% elif is_obs %} 
                      <textarea class="form-input obs-textarea" rows="3" data-pidx="{{ pidx }}" name="field__{{ fidx }}__{{ pidx }}">{{ f.value }}</textarea>
                      <input type="hidden" name="meta_label__{{ fidx }}__{{ pidx }}" value="{{ f.label }}">
                      <input type="hidden" name="meta_cell__{{ fidx }}__{{ pidx }}" value="{{ f.cell }}">
                      <div class="actions" style="margin-top:6px;">
                        <button class="btn-link" type="button" onclick="clearField(this)">Vaciar</button>
                      </div>

                    {% elif is_torn_seg %}
                      {% set pre_nums = p.torn_obs_nums or [] %}
                      {% set base_num = pre_nums[0] if pre_nums|length > 0 else '' %}

                      <div class="inline-compact">
                        <input class="form-input mini-input torn-segons" type="text" inputmode="decimal"
                               name="field__{{ fidx }}__{{ pidx }}" value="{{ f.value }}" autocomplete="off">
                        <select class="form-select mini-select torn-maq-base" name="torn_maq_base__{{ pidx }}" data-selected="{{ base_num|e }}">
                          <option value="">Màq</option>
                        </select>
                        <input type="hidden" name="meta_label__{{ fidx }}__{{ pidx }}" value="{{ f.label }}">
                        <input type="hidden" name="meta_cell__{{ fidx }}__{{ pidx }}" value="{{ f.cell }}">
                      </div>

                      <div id="torn-extra-wrap-{{ pidx }}" style="margin-top:8px;">
                        {% set extra_times = p.torn_extra_times or [] %}
                        {% for t in extra_times %}
                          {% set nsel = pre_nums[loop.index] if pre_nums|length > loop.index else '' %}
                          <div class="inline-compact" style="margin-top:6px;">
                            <input class="form-input mini-input" type="text" inputmode="decimal"
                                   name="torn_time__{{ pidx }}__{{ loop.index }}" value="{{ t|e }}" autocomplete="off" placeholder="Segons/peça">
                            <select class="form-select mini-select torn-maq-extra" name="torn_maq__{{ pidx }}__{{ loop.index }}" data-selected="{{ nsel|e }}">
                              <option value="">Màq</option>
                            </select>
                            <button class="btn-link" type="button" onclick="this.parentElement.remove()">Eliminar</button>
                          </div>
                        {% endfor %}
                      </div>
                      <div class="actions" style="margin-top:6px;">
                        <button class="btn" type="button" onclick="addTornRow({{ pidx }})">Añadir temps cicle</button>
                        <button class="btn-link" type="button" onclick="clearTornRows({{ pidx }})">Vaciar añadidos</button>
                      </div>
                      <input type="hidden" name="torn_extra_count__{{ pidx }}" id="torn_extra_count__{{ pidx }}" value="{{ (p.torn_extra_times|length) if p.torn_extra_times else 0 }}">

                    {% else %}
                      <input class="form-input {% if p.phase_key == 'Neteja' and f.label == 'Temps cicle (min)' %}neteja-tcycle{% endif %}" 
                             type="text" {% if is_numeric %}inputmode="decimal"{% endif %}
                             name="field__{{ fidx }}__{{ pidx }}" value="{{ f.value }}" autocomplete="off" data-pidx="{{ pidx }}">
                      <input type="hidden" name="meta_label__{{ fidx }}__{{ pidx }}" value="{{ f.label }}">
                      <input type="hidden" name="meta_cell__{{ fidx }}__{{ pidx }}" value="{{ f.cell }}">
                      <div class="actions" style="margin-top:6px;">
                        <button class="btn-link" type="button" onclick="clearField(this)">Vaciar</button>
                      </div>
                    {% endif %}
                  </div>
                {% endfor %}
              </div>
            </div>
          {% endfor %}

          <div class="sticky-footer" style="margin-top:6px;">
            <div class="actions">
              <button class="btn" type="submit">Guardar (sobrescribir en PC)</button>
              <a class="btn-link" href="{{ url_for('index') }}">Volver</a>
            </div>
          </div>
        </form>
      {% endif %}
    </main>

    <div class="gyro-wrap" aria-hidden="true" style="position:fixed; right:16px; bottom:16px;">
      <div class="gyro">
        <div class="ring x"></div>
        <div class="ring y"></div>
        <div class="ring z"></div>
      </div>
    </div>

    <script>
      (function () {
        const overlay = document.getElementById('loading-overlay');
        const form = document.getElementById('edit-form');
        if (!overlay || !form) return;
        form.addEventListener('submit', function () {
          overlay.style.display = 'flex';
        });
        window.addEventListener('pagehide', () => {
          overlay.style.display = 'none';
        }, { once: true });
      })();

      function clearField(btn){
        const group = btn.closest('[data-field-group]');
        if(!group) return;
        const ctl = group.querySelector('input.form-input, textarea.form-input');
        if(ctl){
          ctl.value = '';
          ctl.dispatchEvent(new Event('input', { bubbles: true }));
          ctl.dispatchEvent(new Event('change', { bubbles: true }));
          ctl.focus();
        }
      }

      function ccVaciar(btn){
        const group = btn.closest('[data-field-group]');
        if (!group) return;
        const tipus = group.querySelector('.cc-tipus');
        const maq   = group.querySelector('.cc-maquina');
        const code  = group.querySelector('.cc-code');
        const num   = group.parentElement.querySelector('.cc-num');
        if (tipus) tipus.value = '';
        if (maq)   maq.innerHTML = '<option value="">-- Màquina --</option>';
        if (num)   num.innerHTML = '<option value="">-- Nº Màquina (col D) — por E --</option>';
        if (code)  code.value = '';
      }

      // ====== CC widgets (Tipus -> Màquina (dedupe) -> Nº Màquina por E) + Neteja MATRIX ======
      (async function(){
        async function initCentreCostWidgets(){
          try{
            const resp = await fetch("{{ url_for('cc_data') }}");
            const data = await resp.json();
            if (!data || !data.ok) return;

            const types   = data.types || [];
            const items   = data.items_by_type || {};
            const dePairs = data.de_pairs || [];
            const neteja  = (data.neteja_matrix || {machines:[], programs:[], values:{}});

            function normJS(s){
              return (s || '')
                .normalize('NFD').replace(/[\\u0300-\\u036f]/g,'')
                .toLowerCase()
                .replace(/[:()/%]/g, ' ')
                .replace(/[^a-z0-9./\\s-]/g,' ')
                .replace(/\\s+/g,' ')
                .trim();
            }
            function fuzzyMatch(hay, needle){
              const H = normJS(hay);
              const N = normJS(needle);
              if (!N) return false;
              if (H.includes(N) || N.includes(H)) return true;
              const toks = N.split(' ').filter(Boolean);
              if (!toks.length) return false;
              let hit = 0;
              for (const t of toks){ if (H.includes(t)) hit++; }
              return hit >= Math.ceil(toks.length * 0.5);
            }

            // ====== CC (mostrar en fases que no sean Neteja/Centres exteriors) ======
            document.querySelectorAll('.cc-code').forEach(function(hidden){
              const group = hidden.closest('[data-field-group]');
              const card  = hidden.closest('.phase-card');
              const tipusSel = group.querySelector('select.cc-tipus');
              const maqSel   = group.querySelector('select.cc-maquina');
              const numSel   = group.parentElement.querySelector('select.cc-num');

              if (tipusSel){
                tipusSel.innerHTML = '<option value="">-- Tipus --</option>' +
                  types.map(t => `<option value="${t}">${t || '(Sense tipus)'}</option>`).join('');
                tipusSel.addEventListener('change', function(){
                  fillMaquines(this.value, null);
                  hidden.value = "";
                });
              }

              function fillMaquines(tipus, preCode){
                const list = items[tipus] || [];
                const seen = new Set();
                const opts = [];
                for (const it of list){
                  const key = normJS(it.desc);
                  if (seen.has(key)) continue;
                  seen.add(key);
                  opts.push(`<option value="${it.code}">${it.desc}</option>`);
                }
                if (maqSel){
                  maqSel.innerHTML = '<option value="">-- Màquina --</option>' + opts.join('');
                  if (preCode){ maqSel.value = preCode; }
                }
                if (numSel) numSel.innerHTML = '<option value="">-- Nº Màquina (col D) — por E --</option>';
              }

              if (maqSel){
                maqSel.addEventListener('change', function(){
                  hidden.value = this.value || "";
                  if (numSel){
                    const selectedText = this.options[this.selectedIndex] ? this.options[this.selectedIndex].textContent : '';
                    const seen = new Set();
                    const nums = [];
                    for (const p of dePairs){
                      const eText = p && p.text ? p.text : '';
                      const dNum  = p && p.num  ? p.num  : '';
                      if (!dNum) continue;
                      if (fuzzyMatch(eText, selectedText) || fuzzyMatch(selectedText, eText)){
                        if (!seen.has(dNum)){
                          seen.add(dNum); nums.push(dNum);
                        }
                      }
                    }
                    nums.sort((a,b)=> String(a).localeCompare(String(b), undefined, {numeric:true, sensitivity:'base'}));
                    numSel.innerHTML = '<option value="">-- Nº Màquina (col D) — por E --</option>' +
                      nums.map(n => `<option value="${n}">${n}</option>`).join('');
                  }

                  // Torn: al seleccionar CC, sello de tiempo
                  const phaseName = (card.querySelector('.badge')?.textContent || '').trim();
                  if (phaseName === 'Torn'){
                    updateTornChronoStamp(card);
                  }
                });
              }

              // Preselección si ya hay código guardado
              const preCode = hidden.value;
              if (preCode && tipusSel){
                let foundTipus = "";
                for (const t of types){
                  const list = items[t] || [];
                  if (list.some(x => x.code === preCode)){ foundTipus = t; break; }
                }
                if (foundTipus){
                  tipusSel.value = foundTipus;
                  fillMaquines(foundTipus, preCode);
                  if (maqSel) maqSel.dispatchEvent(new Event('change', {bubbles:true}));
                }
              }
            });

            // ====== Mini desplegables de Torn ======
            const allNumsSet = new Set();
            for (const p of dePairs){ if (p && p.num) allNumsSet.add(String(p.num)); }
            const allNums = Array.from(allNumsSet).sort((a,b)=> String(a).localeCompare(String(b), undefined, {numeric:true, sensitivity:'base'}));

            function buildOptionsFromContext(cardRef){
              const ccMaq = cardRef ? cardRef.querySelector('.cc-maquina') : null;
              const ccText = ccMaq && ccMaq.options && ccMaq.selectedIndex > -1
                ? (ccMaq.options[ccMaq.selectedIndex].textContent || "").trim()
                : "";
              if (!ccText){
                return ['<option value="">Màq</option>'].concat(allNums.map(n=>`<option value="${n}">${n}</option>`)).join('');
              }
              const seen = new Set();
              const nums = [];
              for (const p of dePairs){
                const eText = p && p.text ? p.text : '';
                const dNum  = p && p.num  ? String(p.num) : '';
                if (!dNum) continue;
                if (fuzzyMatch(eText, ccText) || fuzzyMatch(ccText, eText)){
                  if (!seen.has(dNum)){ seen.add(dNum); nums.push(dNum); }
                }
              }
              nums.sort((a,b)=> String(a).localeCompare(String(b), undefined, {numeric:true, sensitivity:'base'}));
              const list = nums.length ? nums : allNums;
              return ['<option value="">Màq</option>'].concat(list.map(n=>`<option value="${n}">${n}</option>`)).join('');
            }

            document.querySelectorAll('.phase-card').forEach(card=>{
              card.querySelectorAll('select.torn-maq-base').forEach(sel=>{
                sel.innerHTML = buildOptionsFromContext(card);
                const want = sel.getAttribute('data-selected') || '';
                if (want) sel.value = want;
                sel.addEventListener('change', ()=> updateTornObsInCard(card));
              });
              card.querySelectorAll('select.torn-maq-extra').forEach(sel=>{
                sel.innerHTML = buildOptionsFromContext(card);
                const want = sel.getAttribute('data-selected') || '';
                if (want) sel.value = want;
                sel.addEventListener('change', ()=> updateTornObsInCard(card));
              });

              const ccMaq = card.querySelector('.cc-maquina');
              if (ccMaq){
                ccMaq.addEventListener('change', ()=>{
                  const opts = buildOptionsFromContext(card);
                  card.querySelectorAll('select.torn-maq-base, select.torn-maq-extra').forEach(s=>{
                    const prev = s.value;
                    s.innerHTML = opts;
                    if ([...s.options].some(o=>o.value===prev)) s.value = prev;
                  });
                  updateTornObsInCard(card);
                });
              }
            });
            window.__buildTornOptions = buildOptionsFromContext;

            // ====== Neteja: poblar selects y autocompletar Temps cicle (min), preservando selección del usuario ======
            function updateNetejaTempsFor(pidx){
              const card = document.querySelector(`[name="neteja_maq__${pidx}"]`)?.closest('.phase-card');
              if (!card) return;
              const maqSel = card.querySelector(`[name="neteja_maq__${pidx}"]`);
              const progSel = card.querySelector(`[name="neteja_prog__${pidx}"]`);
              const tcycle = card.querySelector('.neteja-tcycle');
              if (!maqSel || !progSel || !tcycle) return;
              const maq = maqSel.value || '';
              const prog = progSel.value || '';
              let val = null;
              if (maq && prog && neteja.values && neteja.values[maq] && (prog in neteja.values[maq])) {
                val = neteja.values[maq][prog];
              }
              tcycle.value = (val !== null && val !== undefined) ? String(val) : '';
              tcycle.dispatchEvent(new Event('input', {bubbles:true}));
              tcycle.dispatchEvent(new Event('change', {bubbles:true}));
            }
            function fillNetejaForPidx(pidx){
              const maqSel = document.querySelector(`[name="neteja_maq__${pidx}"]`);
              const progSel = document.querySelector(`[name="neteja_prog__${pidx}"]`);
              if (!maqSel || !progSel) return;
              const wantM = maqSel.getAttribute('data-selected') || '';
              const wantP = progSel.getAttribute('data-selected') || '';
              maqSel.innerHTML = '<option value="">-- Selecciona --</option>' + (neteja.machines||[]).map(m=>`<option value="${m}">${m}</option>`).join('');
              progSel.innerHTML = '<option value="">-- Selecciona --</option>' + (neteja.programs||[]).map(p=>`<option value="${p}">${p}</option>`).join('');
              if (wantM) maqSel.value = wantM;
              if (wantP) progSel.value = wantP;
              maqSel.addEventListener('change', ()=>{
                updateNetejaTempsFor(pidx);
                updateNetejaObsFor(pidx);
              });
              progSel.addEventListener('change', ()=>{
                updateNetejaTempsFor(pidx);
                updateNetejaObsFor(pidx);
              });
              // NUEVO: Nau refresca Observacions al instante
              const nauSel = document.querySelector(`[name="neteja_nau__${pidx}"]`);
              if (nauSel) nauSel.addEventListener('change', ()=> updateNetejaObsFor(pidx));

              updateNetejaTempsFor(pidx);
            }
            document.querySelectorAll('[name^="neteja_maq__"]').forEach(el=>{
              const m = el.name.match(/neteja_maq__([0-9]+)/);
              if (m) fillNetejaForPidx(m[1]);
            });

          }catch(e){
            console.warn("CC/NETEJA init error:", e);
          }
        }
        if (document.querySelector('.cc-code') || document.querySelector('.torn-maq-base') || document.querySelector('.neteja-maq')) initCentreCostWidgets();
      })();

      // ====== Torn dinámico (añadir filas) ======
      function addTornRow(pidx){
        const wrap = document.getElementById('torn-extra-wrap-' + pidx);
        const counter = document.getElementById('torn_extra_count__' + pidx);
        if (!wrap || !counter) return;
        const idx = parseInt(counter.value || '0', 10) + 1;
        counter.value = String(idx);
        const card = wrap.closest('.phase-card');
        const selOptions = (window.__buildTornOptions ? window.__buildTornOptions(card) : '<option value="">Màq</option>');
        const row = document.createElement('div');
        row.className = 'inline-compact';
        row.style.marginTop = '6px';
        row.innerHTML = `
          <input class="form-input mini-input" type="text" inputmode="decimal"
                 name="torn_time__${pidx}__${idx}" value="" autocomplete="off" placeholder="Segons/peça">
          <select class="form-select mini-select torn-maq-extra" name="torn_maq__${pidx}__${idx}">
            ${selOptions}
          </select>
          <button class="btn-link" type="button" onclick="this.parentElement.remove(); const card=this.closest('.phase-card'); updateTornObsInCard(card);">Eliminar</button>
        `;
        wrap.appendChild(row);
        const sel = row.querySelector('select.torn-maq-extra');
        if (sel){
          sel.addEventListener('change', ()=>{
            const cardNow = sel.closest('.phase-card');
            updateTornObsInCard(cardNow);
          });
        }
      }
      function clearTornRows(pidx){
        const wrap = document.getElementById('torn-extra-wrap-' + pidx);
        const counter = document.getElementById('torn_extra_count__' + pidx);
        if (wrap) wrap.innerHTML = '';
        if (counter) counter.value = '0';
        const card = document.querySelector('#torn-extra-wrap-' + pidx)?.closest('.phase-card');
        if (card) updateTornObsInCard(card);
      }

      // ====== Live Observacions: Neteja ======
      function updateNetejaObsFor(pidx){
        const tip = document.querySelector(`[name="bast_tipus__${pidx}"]`);
        const card = tip ? tip.closest('.phase-card') : null;
        if (!card) return;
        const obs = card.querySelector('.obs-textarea');
        if (!obs) return;

        const tipus = (card.querySelector(`[name="bast_tipus__${pidx}"]`)?.value || '').trim();
        const maq   = (card.querySelector(`[name="neteja_maq__${pidx}"]`)?.value || '').trim();
        const prog  = (card.querySelector(`[name="neteja_prog__${pidx}"]`)?.value || '').trim();
        const nau   = (card.querySelector(`[name="neteja_nau__${pidx}"]`)?.value || '').trim();

        const parts = [];
        if (tipus) parts.push(`Bastidors: ${tipus}`);
        if (maq)   parts.push(`Màq ${maq}`);
        if (prog)  parts.push(`Programa ${prog}`);
        if (nau)   parts.push(`Nau ${nau}`);

        const line = parts.join(', ');
        const lines = (obs.value || '').split(/\\r?\\n/);
        const filtered = lines.filter(l => !(/(^|\\s)(Bastidors:|Màq\\s|Programa\\s|Nau\\s)/i.test(l.trim())));
        if (line) filtered.push(line);
        obs.value = filtered.join('\\n');
        obs.dispatchEvent(new Event('input', {bubbles:true}));
      }

      // ====== Live Observacions: Torn ======
      function updateTornObsInCard(card){
        if (!card) return;
        const obs = card.querySelector('.obs-textarea');
        if (!obs) return;

        // Recoger todas las máquinas (base + extras)
        const nums = [];
        const base = card.querySelector('select.torn-maq-base');
        const bval = base ? (base.value || '').trim() : '';
        if (bval) nums.push(bval);
        card.querySelectorAll('select.torn-maq-extra').forEach(s=>{
          const v = (s.value || '').trim();
          if (v) nums.push(v);
        });

        // Dedupe manteniendo orden
        const uniq = [];
        const seen = new Set();
        nums.forEach(n=>{ if(!seen.has(n)){ seen.add(n); uniq.push(n);} });

        // Si no hay selección no tocamos Observacions (así no se borra nada)
        if (!uniq.length) return;

        const line = 'Torns: ' + uniq.map(n=>'torn '+n).join(', ');

        const lines = (obs.value || '').split(/\\r?\\n/);
        let replaced = false;
        const out = lines.map(l=>{
          if (/^\\s*Torns\\s*:/i.test(l.trim())){
            replaced = true;
            return line;
          }
          return l;
        });
        if (!replaced) out.push(line);

        // Escribimos con salto de línea real
        obs.value = out.join('\\n');
        obs.dispatchEvent(new Event('input', {bubbles:true}));
      }

      // ====== Torn: sello de tiempo ======
      function pad2(n){ return (n<10?'0':'') + n; }
      function nowStamp(){
        const d = new Date();
        return pad2(d.getDate()) + '/' + pad2(d.getMonth()+1) + '/' + d.getFullYear() + ' ' + pad2(d.getHours()) + ':' + pad2(d.getMinutes());
      }
      function updateTornChronoStamp(card){
        if (!card) return;
        const obs = card.querySelector('.obs-textarea');
        if (!obs) return;
        const stampLine = "Temps cronometrats a " + nowStamp();
        const lines = (obs.value || '').split(/\\r?\\n/);
        let replaced = false;
        const out = lines.map(l=>{
          if (/^\\s*temps\\s+cronometrats\\s+a\\s+/i.test(l.trim())){
            replaced = true;
            return stampLine;
          }
          return l;
        });
        if (!replaced) out.push(stampLine);
        obs.value = out.join('\\n');
        obs.dispatchEvent(new Event('input', {bubbles:true}));
      }

      function initLiveTorn(){
        document.querySelectorAll('.phase-card').forEach(card=>{
          const base = card.querySelector('select.torn-maq-base');
          if (base){
            base.addEventListener('change', ()=> updateTornObsInCard(card));
          }
          card.querySelectorAll('select.torn-maq-extra').forEach(s=>{
            s.addEventListener('change', ()=> updateTornObsInCard(card));
          });
          const segInput = card.querySelector('input.torn-segons');
          if (segInput){
            segInput.addEventListener('input', ()=>{
              if ((segInput.value || '').trim() !== ''){
                updateTornChronoStamp(card);
              }
            });
          }
        });
      }

      // NUEVO: inicialización de listeners Neteja (para refrescar Observacions)
      function initLiveNeteja(){
        document.querySelectorAll('[name^="bast_tipus__"]').forEach(input=>{
          const m = input.name.match(/bast_tipus__([0-9]+)/);
          if (!m) return;
          const pidx = m[1];
          ['bast_tipus__','bast_num__','bast_ppb__','neteja_nau__'].forEach(prefix=>{
            const ctl = document.querySelector(`[name="${prefix}${pidx}"]`);
            if (ctl){
              const evt = prefix === 'neteja_nau__' ? 'change' : 'input';
              ctl.addEventListener(evt, ()=> updateNetejaObsFor(pidx));
            }
          });
          // Inicial
          updateNetejaObsFor(pidx);
        });
      }

      document.addEventListener('DOMContentLoaded', ()=>{
        initLiveNeteja();
        initLiveTorn();
      });

      lucide.createIcons();
    </script>
  </body>
</html>
"""

# ========= Rutas =========
@app.route("/", methods=["GET"])
def index():
    return render_template_string(
        INDEX_HTML,
        files=listar_excels_recursivo(BASE_TEMPLATE_DIR),
        base_dir=BASE_TEMPLATE_DIR,
        msg=None,
        BASE_CSS=BASE_CSS
    )

@app.route("/cc_data", methods=["GET"])
def cc_data():
    data = load_cc_table()
    return {
        "ok": True,
        "path": CC_XLSX_PATH,
        "types": data["types"],
        "items_by_type": data["items_by_type"],
        "de_pairs": data["de_pairs"],
        "neteja_matrix": data.get("neteja_matrix", {"machines":[], "programs":[], "values":{}}),
    }

def _sticky_from_form(frm):
    """Persistencia en UI tras guardar: solo inputs de Neteja que no se guardan en Excel."""
    keep = {}
    for k in frm.keys():
        if k.startswith(("bast_tipus__", "bast_num__", "bast_ppb__", "neteja_maq__", "neteja_prog__", "neteja_nau__")):
            keep[k] = frm.get(k, "")
    return keep

@app.route("/edit", methods=["GET", "POST"])
def edit():
    if request.method == "GET":
        relpath = (request.args.get("file") or "").strip()
        if not relpath:
            return redirect(url_for("index"))
        abs_path = os.path.realpath(os.path.join(BASE_TEMPLATE_DIR, relpath.replace("/", os.sep)))
        if not abs_path.startswith(os.path.realpath(BASE_TEMPLATE_DIR)) or not os.path.isfile(abs_path):
            return render_template_string(INDEX_HTML, files=listar_excels_recursivo(BASE_TEMPLATE_DIR), base_dir=BASE_TEMPLATE_DIR, msg="Ruta no permitida o archivo no existe.", BASE_CSS=BASE_CSS)
        debug = request.args.get("debug") in ("1", "true", "yes")
        schema = detect_schema(abs_path, debug=debug)
        return render_template_string(
            EDIT_HTML,
            relpath=relpath,
            piece_name=piece_name_from_relpath(relpath),
            schema=schema,
            BASE_CSS=BASE_CSS,
            sticky={}  # GET: sin pegajosos
        )

    # POST -> Guardar
    relpath = (request.form.get("file") or "").strip()
    abs_path = os.path.realpath(os.path.join(BASE_TEMPLATE_DIR, relpath.replace("/", os.sep)))
    if not abs_path.startswith(os.path.realpath(BASE_TEMPLATE_DIR)) or not os.path.isfile(abs_path):
        return render_template_string(INDEX_HTML, files=listar_excels_recursivo(BASE_TEMPLATE_DIR), base_dir=BASE_TEMPLATE_DIR, msg="Ruta no permitida o archivo no existe.", BASE_CSS=BASE_CSS)

    def to_colrow(a1: str):
        m = re.match(r"([A-Za-z]+)(\d+)$", a1 or "")
        if not m: return None, None
        col = column_index_from_string(m.group(1))
        row = int(m.group(2))
        return col, row

    def do_write():
        wb = load_workbook(abs_path)
        ws = wb[wb.sheetnames[0]]
        merge_map = build_merge_map(ws)

        # Escritura de campos simples
        for k in list(request.form.keys()):
            if not k.startswith("meta_label__"):
                continue
            try:
                _, idxs = k.split("meta_label__", 1)
                idxs = idxs.strip("_")
                f_idx, p_idx = idxs.split("__")
            except Exception:
                continue

            cell_ref = request.form.get(f"meta_cell__{f_idx}__{p_idx}")
            if cell_ref is None:
                continue
            val = request.form.get(f"field__{f_idx}__{p_idx}", "")

            if not cell_ref:
                continue

            m = re.match(r"([A-Z]+)(\d+)", cell_ref, re.I)
            if not m:
                continue
            col_str, row_str = m.group(1).upper(), m.group(2)
            col_num = 0
            for ch in col_str:
                col_num = col_num * 26 + (ord(ch) - ord('A') + 1)
            row_num = int(row_str)

            write_ref = cell_ref_for_write_cached(merge_map, row_num, col_num)
            try:
                ws[write_ref].value = (None if val == "" else val)
            except Exception:
                pass

        # ====== Neteja: calcular Quantitat y Observacions ======
        pidx_set = set()
        for k in list(request.form.keys()):
            if k.startswith("meta_quant_cell__"):
                m = re.match(r"meta_quant_cell__([0-9]+)$", k)
                if m: pidx_set.add(m.group(1))

        for pidx in pidx_set:
            quant_cell = request.form.get(f"meta_quant_cell__{pidx}", "")
            obs_cell   = request.form.get(f"meta_obs_cell__{pidx}", "")

            # qty = num * ppb
            try:
                num_v = float(request.form.get(f"bast_num__{pidx}", "") or 0)
                ppb_v = float(request.form.get(f"bast_ppb__{pidx}", "") or 0)
                qty = num_v * ppb_v
            except Exception:
                qty = None

            # escribir Quantitat
            c, r = to_colrow(quant_cell)
            if c and r:
                try:
                    ws[cell_ref_for_write_cached(merge_map, r, c)].value = (None if qty is None else qty)
                except Exception:
                    pass

            # Observacions en línea con comas + sello si no existe
            if obs_cell:
                oc, orow = to_colrow(obs_cell)
                if oc and orow:
                    ref = cell_ref_for_write_cached(merge_map, orow, oc)
                    cur = ws[ref].value
                    cur_s = "" if cur is None else str(cur)

                    parts = []
                    tipus = (request.form.get(f"bast_tipus__{pidx}", "") or "").strip()
                    if tipus: parts.append(f"Bastidors: {tipus}")
                    maq = (request.form.get(f"neteja_maq__{pidx}", "") or "").strip()
                    if maq: parts.append(f"Màq {maq}")
                    prog = (request.form.get(f"neteja_prog__{pidx}", "") or "").strip()
                    if prog: parts.append(f"Programa {prog}")
                    nau = (request.form.get(f"neteja_nau__{pidx}", "") or "").strip()
                    if nau: parts.append(f"Nau {nau}")

                    if parts:
                        lines = cur_s.splitlines()
                        lines = [l for l in lines if not re.search(r'(^|\\s)(Bastidors:|Màq\\s|Programa\\s|Nau\\s)', l, flags=re.I)]
                        lines.append(", ".join(parts))
                        cur_s = "\n".join(lines)

                    stamp = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                    if "temps cronometrats a" not in norm_txt(cur_s):
                        if cur_s and not cur_s.endswith("\n"): cur_s += "\n"
                        cur_s += f"Temps cronometrats a {stamp}"

                    ws[ref].value = cur_s or None

        # ====== Torn: extras debajo + Observacions "Torns: ..." ======
        torn_pidx_to_cell = {}
        torn_pidx_to_obs  = {}
        for k in list(request.form.keys()):
            if k.startswith("meta_label__"):
                try:
                    _, idxs = k.split("meta_label__", 1)
                    idxs = idxs.strip("_")
                    f_idx, p_idx = idxs.split("__")
                except Exception:
                    continue
                label = request.form.get(k, "")
                if label == "Segons/peça":
                    cell_ref = request.form.get(f"meta_cell__{f_idx}__{p_idx}", "")
                    if cell_ref:
                        c, r = to_colrow(cell_ref)
                        if c and r:
                            torn_pidx_to_cell[p_idx] = (c, r)

        for k in list(request.form.keys()):
            m = re.match(r"meta_obs_cell__([0-9]+)$", k)
            if m:
                pidx = m.group(1)
                refs = request.form.get(k, "")
                if refs:
                    c, r = to_colrow(refs)
                    if c and r:
                        torn_pidx_to_obs[pidx] = (c, r)

        for pidx, base in torn_pidx_to_cell.items():
            base_c, base_r = base
            try:
                extra_n = int(request.form.get(f"torn_extra_count__{pidx}", "0") or "0")
            except Exception:
                extra_n = 0

            # escribir extras
            for i in range(1, extra_n + 1):
                val = (request.form.get(f"torn_time__{pidx}__{i}", "") or "").strip()
                if val == "":
                    continue
                dest_ref = cell_ref_for_write_cached(merge_map, base_r + i, base_c)
                try:
                    ws[dest_ref].value = val
                except Exception:
                    pass

            # Observacions: Torns
            # Observacions: Torns
            maq_nums = []
            base_maq = (request.form.get(f"torn_maq_base__{pidx}", "") or "").strip()
            if base_maq:
                maq_nums.append(base_maq)
            for i in range(1, extra_n + 1):
                mv = (request.form.get(f"torn_maq__{pidx}__{i}", "") or "").strip()
                if mv:
                    maq_nums.append(mv)

            if maq_nums and (pidx in torn_pidx_to_obs):
                seen = set()
                maq_final = []
                for n in maq_nums:
                    if n not in seen:
                        seen.add(n);
                        maq_final.append(n)
                line = "Torns: " + ", ".join([f"torn {n}" for n in maq_final])

                c, r = torn_pidx_to_obs[pidx]
                ref = cell_ref_for_write_cached(merge_map, r, c)
                cur = ws[ref].value
                cur_s = "" if cur is None else str(cur)

                # 👉 Solo añadir si NO hay ninguna línea "Torns:" ya
                if not re.search(r'(^|\n)\s*Torns\s*:', cur_s, flags=re.I):
                    if cur_s and not cur_s.endswith("\n"):
                        cur_s += "\n"
                    cur_s += line

                ws[ref].value = cur_s or None

        wb.save(abs_path)
        wb.close()

    if HAVE_PORTALOCKER:
        try:
            with portalocker.Lock(abs_path, timeout=10, flags=portalocker.LOCK_EX):
                do_write()
        except Exception:
            do_write()
    else:
        do_write()

    schema = detect_schema(abs_path)
    return render_template_string(
        EDIT_HTML,
        relpath=relpath,
        piece_name=piece_name_from_relpath(relpath),
        schema=schema,
        BASE_CSS=BASE_CSS,
        sticky=_sticky_from_form(request.form)  # mantener inputs de Neteja tras guardar
    )

# (opcional) cabeceras mínimas + CSP
@app.after_request
def add_security_headers(resp):
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("Referrer-Policy", "same-origin")
    resp.headers.setdefault("X-Frame-Options", "DENY")
    resp.headers.setdefault("Content-Security-Policy",
        "default-src 'self'; "
        "img-src 'self' data:; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "script-src 'self' 'unsafe-inline' https://unpkg.com; "
    )
    return resp

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)
